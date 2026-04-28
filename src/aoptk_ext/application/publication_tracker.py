# ruff: noqa: PLR0913
from __future__ import annotations
import time
from pathlib import Path
import click
import pandas as pd
from aoptk.literature.databases.europepmc import EuropePMC
from aoptk.literature.databases.pubmed import PubMed
from Bio import Entrez
from openpyxl import load_workbook


@click.command()
@click.option(
    "--read",
    type=str,
    required=True,
    help="Provide path to the database of read publications",
)
@click.option("--master", type=str, required=True, help="Provide path to the master table")
@click.option("--code", type=str, required=True, help="Provide search code to track this search")
@click.option("--email", type=str, required=False, help="Email address to follow PubMed - NCBI guidelines")
@click.option(
    "--query",
    type=str,
    required=True,
    help="Search term for PubMed or Europe PMC",
)
@click.option(
    "--database",
    type=click.Choice(["pubmed", "europepmc"]),
    required=True,
    help="Database to search: PubMed or Europe PMC",
)
@click.option("--outdir", "-o", type=str, required=True, help="Output directory.")
def cli(read: str, master: str, email: str, code: str, query: str, database: str, outdir: str) -> None:
    """Generate publications to read and update master table search codes."""
    if database == "pubmed":
        Entrez.email = email
        db = PubMed(query)
    elif database == "europepmc":
        db = EuropePMC(query)
    else:
        msg = "Unknown database."
        raise ValueError(msg)

    metadata = convert_metadata_structures_to_df(
        code,
        query,
        database,
        db.get_publications_metadata(),
    )

    generate_publications_to_read(read, metadata, outdir)
    update_master_table_search_codes(master, code, metadata, outdir)


def convert_metadata_structures_to_df(
    search_code: str,
    query: str,
    literature_database: str,
    publications_metadata: list,
) -> pd.DataFrame:
    """Convert list of publication metadata structures to dataframe."""
    rows = [
        {
            "id": pub.publication_id,
            "year_publication": pub.publication_date,
            "authors": pub.authors,
            "title": pub.title,
            "search_term": query,
            "search_code": search_code,
            "search_date": time.strftime("%Y-%m-%d"),
            "database": literature_database,
        }
        for pub in publications_metadata
    ]
    return pd.DataFrame(rows)


def generate_publications_to_read(database_path: str, metadata: pd.DataFrame, outdir: str) -> None:
    """Generate publications to read based on existing database of read publications."""
    read_publications = pd.read_excel(database_path)
    existing_ids = read_publications["id"].dropna().astype(str)
    to_read = metadata.loc[~(metadata["id"].isin(existing_ids))]
    to_read.to_excel(Path(outdir) / "to_read.xlsx", index=False)


def update_master_table_search_codes(
    master_table_path: str,
    search_code: str,
    metadata_df: pd.DataFrame,
    outdir: str,
) -> None:
    """Update master table with new search codes."""
    master_wb = load_workbook(master_table_path)
    master_ws = master_wb.active
    header = [cell.value for cell in master_ws[1]]
    master_id_col = header.index("ID")
    master_search_code = header.index("Search code")
    master_id_map = create_map_of_ids_from_master_table(master_ws, master_id_col)
    publications_id = metadata_df["id"].astype(str)
    common_ids_to_read_publications_master = set(publications_id).intersection(master_id_map.keys())
    for row in metadata_df.itertuples(index=False):
        row_id = str(row.id)
        if row_id in common_ids_to_read_publications_master:
            for excel_row_idx in master_id_map[row_id]:
                cell = master_ws.cell(row=excel_row_idx, column=master_search_code + 1)
                current_value = cell.value
                updated_value = f"{current_value} ; {search_code}" if current_value else search_code
                cell.value = updated_value
    master_wb.save(Path(outdir) / "updated_master_table.xlsx")


def create_map_of_ids_from_master_table(master_ws: object, master_id_col: int) -> dict[str, list[int]]:
    """Create a map of IDs from the master table to their corresponding row indices."""
    master_id_map = {}
    for idx, row in enumerate(master_ws.iter_rows(min_row=2, values_only=True), start=2):
        row_id = str(row[master_id_col]) if row[master_id_col] else None
        if row_id:
            master_id_map.setdefault(row_id, []).append(idx)
    return master_id_map
